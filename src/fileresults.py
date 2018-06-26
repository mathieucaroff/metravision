from pathlib import Path

from openpyxl import Workbook

from datetime import datetime


def fillPathTemplate(videoPath: Path, ext: str, pathTemplate: str, segmentIndexList):
    """
    Generate a filename from a template.

    Get all informations usable in the path template and interpolate them.
    
    ---

    Génère un nom de fichier à partir d'un template.

    Récupère et interpôle toutes les informations utilisables par le template
    `pathTemplate` pour générer un nom de fichier. Cette fonction était
    prévue pour générer le nom du fichier de résultat pour les données extraites
    d'une vidéo.

    :param videoPath Path: Chemin de la video
    :param ext: Extension du fichier de résultat
    :para pathTemplate: Template du fichier de résultat
    :param segmentIndexList: Liste des tranches traitées
    :return: Le nom de fichier à utiliser pour écrire les résultats
    """
    video_directory = videoPath.parent
    video_filename = videoPath.stem
    video_extension = videoPath.suffix[1:]

    now = datetime.now()
    mv_date = now.strftime("%Y-%m-%d")
    mv_datetime = now.strftime("%Y-%m-%dT%H")
    mv_utcdatetime = datetime.utcnow().strftime("%Y-%m-%dT%HZ")
    with Path("VERSION.txt").open() as f:
        mv_version = "-".join(f.readline().split("."))

    # Compute segmentIndex
    if len(segmentIndexList) != 0:
        a = min(segmentIndexList)
        b = max(segmentIndexList)
        if len(segmentIndexList) >= b - a:
            _complete = True
            joiner = "-"
        else:
            _complete = False
            joiner = "--"
        segmentIndex = joiner.join([str(a), str(b)])
    else:
        segmentIndex = "__"

    timedPathTemplate = now.strftime(pathTemplate)
    path = timedPathTemplate.format(
        video_directory = video_directory,
        video_filename = video_filename,
        video_extension = video_extension,
        mv_date = mv_date,
        mv_datetime = mv_datetime,
        mv_utcdatetime = mv_utcdatetime,
        mv_version = mv_version,
        ext = ext,
        segmentIndex = segmentIndex
    )

    return path


def createXLS(sheetContent, sheetHeader=None):
    """
    Convert a list of list of data into a spreadsheet document using openpyxl.

    ---

    Converti une liste de liste de données en un document spreadsheet.

    Utilise pour cela le module openpyxl.
    Le document est renvoyé et peut être enregistré via la méthode 
    `.save(filename)` du document.
    Si un header est donné, les données sont écrite une ligne plus bas.
    """
    wb = Workbook()
    ws = wb.active
    if sheetHeader is not None:
        #for i, val in enumerate(sheetHeader):
        #    ws.cell(row=1, column=i).value = val
        _hasHeader = 1
        sheetContent[:0] = [sheetHeader]
    else:
        _hasHeader = 0
    
    for i, rowContent in enumerate(sheetContent):
        #util.show(rowContent=rowContent)
        for j, val in enumerate(rowContent):
            #util.show(val=val)
            ws.cell(row=i + 1, column=j + 1).value = val
    return wb